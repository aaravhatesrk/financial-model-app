"""
Builds a downloadable .xlsx workbook from the company data + financial model
produced by scraper.get_company() / financial_model.build_model().
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, color="1F3864")
LABEL_FONT = Font(bold=True)


def _autofit(ws, min_width=10, max_width=45):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, min_width), min(length + 2, max_width))
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_row(ws, row_idx, values, bold=False, fill=False):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=i, value=v)
        if bold:
            cell.font = HEADER_FONT if fill else LABEL_FONT
        if fill:
            cell.fill = HEADER_FILL
    return row_idx + 1


def _write_table(ws, start_row, title, periods, rows_dict, row_order=None):
    r = start_row
    ws.cell(row=r, column=1, value=title).font = SUBTITLE_FONT
    r += 1
    keys = list(row_order) if row_order else []
    for k in rows_dict:
        if k not in keys:
            keys.append(k)
    keys = [k for k in keys if k in rows_dict]
    r = _write_row(ws, r, ["Particulars"] + list(periods), bold=True, fill=True)
    for k in keys:
        vals = rows_dict[k]
        padded = list(vals) + [None] * (len(periods) - len(vals))
        r = _write_row(ws, r, [k] + padded)
    return r + 1


def _kv_block(ws, start_row, title, pairs):
    r = start_row
    ws.cell(row=r, column=1, value=title).font = SUBTITLE_FONT
    r += 1
    for label, value in pairs:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r + 1


def build_workbook(data, model):
    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=data.get("name") or "Company").font = TITLE_FONT
    ws.cell(row=2, column=1, value=data.get("url") or "")
    ws.cell(row=3, column=1, value=data.get("about") or "")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    r = 5
    top_ratios = data.get("top_ratios") or {}
    tr_pairs = []
    for name in ["Market Cap", "Current Price", "Stock P/E", "Book Value", "ROCE", "ROE", "Dividend Yield", "Face Value"]:
        entry = top_ratios.get(name)
        tr_pairs.append((name, entry.get("raw") if entry and entry.get("raw") else "N/A"))
    r = _kv_block(ws, r, "Key Ratios (screener.in)", tr_pairs)

    is_bank = model.get("bank_model") is not None
    if is_bank:
        b = model["bank_model"]
        valuation_pairs = [
            ("Model type", "Excess Return / Justified P/B"),
            ("Sustainable ROE (%)", b.get("sustainable_roe_pct")),
            ("Payout ratio (%)", b.get("payout_ratio_pct")),
            ("Sustainable growth, g (%)", b.get("sustainable_growth_pct")),
            ("Cost of equity, Ke (%)", b.get("cost_of_equity_pct")),
            ("Book value / share (Rs)", b.get("book_value_per_share")),
            ("Justified P/B (x)", b.get("justified_pb")),
            ("Fair value / share (Rs)", b.get("value_per_share")),
            ("Current price (Rs)", b.get("current_price")),
            ("Upside / (Downside) (%)", b.get("upside_pct")),
            ("Recommendation", b.get("recommendation")),
        ]
    else:
        d = model.get("dcf") or {}
        valuation_pairs = [
            ("Model type", "10-Year FCFF DCF"),
            ("Enterprise value (Rs Cr)", d.get("enterprise_value")),
            ("Less: Net debt (Rs Cr)", d.get("net_debt")),
            ("Equity value (Rs Cr)", d.get("equity_value")),
            ("Shares outstanding", d.get("shares_outstanding")),
            ("DCF value / share (Rs)", d.get("value_per_share")),
            ("Current price (Rs)", d.get("current_price")),
            ("Upside / (Downside) (%)", d.get("upside_pct")),
            ("Recommendation", d.get("recommendation")),
        ]
    r = _kv_block(ws, r, "Valuation Summary", valuation_pairs)

    comps = model.get("comps") or {}
    comps_pairs = [
        ("Peers found", comps.get("peer_count")),
        ("Peer average P/E (x)", comps.get("peer_avg_pe")),
        ("Subject stock P/E (x)", comps.get("subject_pe")),
        ("Latest EPS (Rs)", comps.get("latest_eps")),
        ("Implied value, peer P/E x EPS (Rs)", comps.get("implied_value_pe")),
        ("Current price (Rs)", comps.get("current_price")),
        ("Upside / (Downside) (%)", comps.get("upside_pct")),
    ]
    r = _kv_block(ws, r, "Relative Valuation (Peer P/E)", comps_pairs)
    _autofit(ws)

    # ---------------- Assumptions ----------------
    ws2 = wb.create_sheet("Assumptions")
    assumptions = model.get("assumptions") or {}
    pairs = []
    for k, v in assumptions.items():
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                pairs.append((f"{k.replace('_', ' ').title()} - {sub_k.replace('_', ' ').title()}", sub_v))
        else:
            pairs.append((k.replace("_", " ").title(), v))
    _kv_block(ws2, 1, "Model Assumptions", pairs)
    _autofit(ws2)

    # ---------------- DCF Projections + Sensitivity ----------------
    if not is_bank and model.get("dcf") and not model["dcf"].get("error"):
        ws3 = wb.create_sheet("DCF Projections")
        projections = model["dcf"]["projections"]
        periods = [f"Yr {p['year']}" for p in projections]
        rows = {
            "Revenue": [p["revenue"] for p in projections],
            "Growth %": [p["growth_pct"] for p in projections],
            "EBITDA": [p["ebitda"] for p in projections],
            "EBIT": [p["ebit"] for p in projections],
            "Tax": [p["tax"] for p in projections],
            "NOPAT": [p["nopat"] for p in projections],
            "D&A": [p["da"] for p in projections],
            "Capex": [p["capex"] for p in projections],
            "Delta NWC": [p["delta_nwc"] for p in projections],
            "FCFF": [p["fcff"] for p in projections],
            "Discount Factor": [p["discount_factor"] for p in projections],
            "PV of FCFF": [p["pv_fcff"] for p in projections],
        }
        _write_table(ws3, 1, "10-Year FCFF Projection (Rs Cr)", periods, rows, list(rows.keys()))
        _autofit(ws3)

        sens = model.get("sensitivity")
        if sens:
            ws4 = wb.create_sheet("Sensitivity")
            ws4.cell(row=1, column=1, value="Value per Share (Rs): WACC (rows) vs Terminal Growth (cols)").font = SUBTITLE_FONT
            r = _write_row(ws4, 2, ["WACC \\ g"] + [f"{g}%" for g in sens["g_range"]], bold=True, fill=True)
            for i, w in enumerate(sens["wacc_range"]):
                r = _write_row(ws4, r, [f"{w}%"] + sens["grid"][i])
            _autofit(ws4)

    if is_bank:
        ws3 = wb.create_sheet("Bank Model")
        b = model["bank_model"]
        pairs = [(k.replace("_", " ").title(), v) for k, v in b.items()]
        _kv_block(ws3, 1, "Excess Return / Justified P/B Model", pairs)
        _autofit(ws3)

    # ---------------- Comps ----------------
    ws5 = wb.create_sheet("Peer Comps")
    peers = comps.get("peers") or []
    if peers:
        headers = list(dict.fromkeys(k for p in peers for k in p.keys()))
        r = _write_row(ws5, 1, headers, bold=True, fill=True)
        for p in peers:
            r = _write_row(ws5, r, [p.get(h) for h in headers])
    else:
        ws5.cell(row=1, column=1, value="No peer data available.")
    _autofit(ws5)

    # ---------------- Historical financials ----------------
    pl = data.get("profit_loss") or {"periods": [], "rows": {}}
    bs = data.get("balance_sheet") or {"periods": [], "rows": {}}
    cf = data.get("cash_flow") or {"periods": [], "rows": {}}
    ratios = data.get("ratios") or {"periods": [], "rows": {}}

    ws6 = wb.create_sheet("Profit & Loss")
    _write_table(ws6, 1, "Profit & Loss (Rs Cr)", pl["periods"], pl["rows"],
                 ["Sales", "Expenses", "Operating Profit", "OPM %", "Other Income", "Interest",
                  "Depreciation", "Profit before tax", "Tax %", "Net Profit", "EPS in Rs", "Dividend Payout %"])
    _autofit(ws6)

    ws7 = wb.create_sheet("Balance Sheet")
    _write_table(ws7, 1, "Balance Sheet (Rs Cr)", bs["periods"], bs["rows"],
                 ["Equity Share Capital", "Reserves", "Borrowings", "Other Liabilities", "Total Liabilities",
                  "Fixed Assets", "CWIP", "Investments", "Other Assets", "Total Assets"])
    _autofit(ws7)

    ws8 = wb.create_sheet("Cash Flow")
    _write_table(ws8, 1, "Cash Flow (Rs Cr)", cf["periods"], cf["rows"],
                 ["Cash from Operating Activity", "Cash from Investing Activity",
                  "Cash from Financing Activity", "Net Cash Flow"])
    _autofit(ws8)

    ws9 = wb.create_sheet("Ratios Trend")
    _write_table(ws9, 1, "Key Ratios Trend", ratios["periods"], ratios["rows"],
                 ["ROCE %", "ROE %", "Debtor Days", "Working Capital Days"])
    _autofit(ws9)

    for ws_ in wb.worksheets:
        ws_.freeze_panes = "A2" if ws_.title in ("Peer Comps",) else None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
